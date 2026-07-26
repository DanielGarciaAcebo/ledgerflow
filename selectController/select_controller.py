from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook


def select_excel_file(
    status_label: ttk.Label,
    header_row: int,
) -> tuple[list[str], list[list[object]]] | None:
    file_path = filedialog.askopenfilename(
        parent=status_label.winfo_toplevel(),
        title="Select Excel File",
        initialdir=str(Path.home()),
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*"),
        ],
    )

    if not file_path:
        status_label.config(text="No file selected")
        return None

    selected_file = Path(file_path)

    try:
        workbook = load_workbook(
            filename=selected_file,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

        header_cells = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )

        headers = [
            str(value).strip() if value is not None else f"Column {index + 1}"
            for index, value in enumerate(header_cells)
        ]

        rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            )
            if any(value is not None for value in row)
        ]

        sheet_name = worksheet.title
        workbook.close()

        status_label.config(
            text=f"Selected: {selected_file.name}",
        )

        print(f"Selected file: {selected_file}")
        print(f"Active sheet: {sheet_name}")
        print(f"Header row: {header_row}")
        print(f"Rows loaded: {len(rows)}")

        return headers, rows

    except Exception as error:
        status_label.config(text="Could not read the file")

        messagebox.showerror(
            title="Excel Error",
            message=f"Could not read the Excel file.\n\n{error}",
        )

        return None