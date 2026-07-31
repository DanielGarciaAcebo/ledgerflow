from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from models.transaction import Transaction
from services.automatic_classifier import parse_amount

def export_transactions_to_excel(
    parent: tk.Misc,
    transactions: list[Transaction],
    groups: list[str],
) -> Path | None:
    if not transactions:
        messagebox.showwarning(
            title="No Transactions",
            message="There are no transactions to export.",
            parent=parent,
        )
        return None

    unclassified_transactions = [
        transaction
        for transaction in transactions
        if not transaction.group_assignments
    ]

    if unclassified_transactions:
        messagebox.showwarning(
            title="Unclassified Transactions",
            message=(
                f"{len(unclassified_transactions)} transactions "
                "have no assigned group.\n\n"
                "Classify them before exporting."
            ),
            parent=parent,
        )
        return None

    output_path = filedialog.asksaveasfilename(
        parent=parent,
        title="Export Excel File",
        initialfile="ledgerflow_output.xlsx",
        defaultextension=".xlsx",
        filetypes=[
            ("Excel files", "*.xlsx"),
        ],
    )

    if not output_path:
        return None

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"

    headers = [
        "Name",
        *groups,
    ]

    worksheet.append(headers)

    # Header formatting
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # Transaction rows
    for transaction in transactions:
        parsed_amount = parse_amount(
            transaction.amount,
        )

        row: list[object] = [
            transaction.name,
        ]

        for group_name in groups:
            if group_name not in transaction.group_assignments:
                row.append(None)
                continue

            if parsed_amount is None:
                row.append(None)
                continue

            invert_sign = transaction.group_assignments[
                group_name
            ]

            final_amount = (
                -parsed_amount
                if invert_sign
                else parsed_amount
            )

            row.append(float(final_amount))

        worksheet.append(row)

    # Excel behavior and formatting
    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions

    worksheet.column_dimensions["A"].width = 45

    for column_index, group_name in enumerate(
        groups,
        start=2,
    ):
        column_letter = get_column_letter(
            column_index,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = max(
            len(group_name) + 4,
            15,
        )

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = '#,##0.00;[Red]-#,##0.00'

    output_file = Path(output_path)

    try:
        workbook.save(output_file)

    except Exception as error:
        messagebox.showerror(
            title="Export Error",
            message=(
                "Could not create the Excel file.\n\n"
                f"{error}"
            ),
            parent=parent,
        )
        return None

    messagebox.showinfo(
        title="Export Complete",
        message=f"Excel file created successfully:\n\n{output_file}",
        parent=parent,
    )

    return output_file