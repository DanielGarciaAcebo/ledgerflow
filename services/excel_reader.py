from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from models.excel_data import ExcelData


class ExcelReadError(Exception):
    """Base exception for Excel reading errors."""


class InvalidHeaderRowError(ExcelReadError):
    """Raised when the selected header row is invalid."""


class HeaderRowNotFoundError(ExcelReadError):
    """Raised when the selected header row does not exist."""


class EmptyExcelFileError(ExcelReadError):
    """Raised when the selected worksheet contains no usable columns."""


def read_excel_file(
    file_path: str | Path,
    header_row: int,
) -> ExcelData:
    selected_file = Path(file_path)

    if header_row < 1:
        raise InvalidHeaderRowError(
            "The header row must be greater than zero."
        )

    if not selected_file.exists():
        raise ExcelReadError(
            f'The file "{selected_file}" does not exist.'
        )

    if not selected_file.is_file():
        raise ExcelReadError(
            f'The path "{selected_file}" is not a file.'
        )

    workbook: Workbook | None = None

    try:
        workbook = load_workbook(
            filename=selected_file,
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

        header_cells = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            ),
            None,
        )

        if header_cells is None:
            raise HeaderRowNotFoundError(
                f"Row {header_row} does not exist in the worksheet."
            )

        if not header_cells:
            raise EmptyExcelFileError(
                "The worksheet does not contain any columns."
            )

        headers = [
            (
                str(value).strip()
                if value is not None
                else f"Column {index + 1}"
            )
            for index, value in enumerate(header_cells)
        ]

        rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            )
            if any(value is not None for value in row)
        ]

        return ExcelData(
            file_path=selected_file,
            sheet_name=worksheet.title,
            headers=headers,
            rows=rows,
        )

    except ExcelReadError:
        raise

    except Exception as error:
        raise ExcelReadError(
            f"Could not read the Excel file: {error}"
        ) from error

    finally:
        if workbook is not None:
            workbook.close()